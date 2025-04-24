import os
import re
import base64
from typing import List, Dict, Optional
from docx import Document
from docx.table import Table
from docx.text.paragraph import Paragraph
from docx.oxml.ns import qn

# Для обработки Excel
from openpyxl import load_workbook

def iter_block_items(parent):
    """
    Генератор для последовательного обхода блоковых элементов (параграфов и таблиц)
    в документе или ячейке таблицы.
    """
    from docx.oxml.text.paragraph import CT_P
    from docx.oxml.table import CT_Tbl

    if hasattr(parent, 'element'):
        parent_elm = parent.element.body
    else:
        parent_elm = parent._tc  # для ячеек таблицы

    for child in parent_elm.iterchildren():
        if child.tag.endswith('}p'):
            yield Paragraph(child, parent)
        elif child.tag.endswith('}tbl'):
            yield Table(child, parent)

def get_image_base64(run):
    for shape in run._element.findall('.//{http://schemas.openxmlformats.org/drawingml/2006/main}blip'):
        rId = shape.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
        image_part = run.part.rels[rId].target_part
        image_data = image_part.blob
        return base64.b64encode(image_data).decode('utf-8')
    return None

def get_paragraph_text_and_images(paragraph):
    """
    Извлекает текст абзаца и отдельно все найденные изображения.
    Если в run встречается картинка (наличие тега "graphic"), то текст из этого run не включается,
    а картинка сохраняется в отдельном списке.
    """
    text_parts = []
    images = []
    for run in paragraph.runs:
        # Если в run присутствует графика – извлекаем картинку и не добавляем её текст в основной контент
        if "graphic" in run._element.xml:
            img_base64 = get_image_base64(run)
            if img_base64:
                images.append(f'<img src="data:image/png;base64,{img_base64}" alt="Картинка">')
        else:
            text_parts.append(run.text)
    return "".join(text_parts).strip(), images

def get_table_html(table: Table) -> str:
    """
    Преобразует таблицу из docx в HTML с корректной обработкой объединённых ячеек.
    Алгоритм строит «сетку» таблицы, в которую заносятся ссылки на верхние левый элементы объединённых областей.
    """
    grid = {}
    num_rows = len(table.rows)
    max_cols = 0

    for r, row in enumerate(table.rows):
        col = 0
        for cell in row.cells:
            while (r, col) in grid:
                col += 1

            tcPr = cell._tc.get_or_add_tcPr()
            colspan = 1
            gridSpan = tcPr.find(qn('w:gridSpan'))
            if gridSpan is not None:
                try:
                    colspan = int(gridSpan.get(qn('w:val')))
                except Exception:
                    colspan = 1

            rowspan = 1
            vMerge = tcPr.find(qn('w:vMerge'))
            if vMerge is not None:
                vMerge_val = vMerge.get(qn('w:val'))
                if vMerge_val is None or vMerge_val == 'restart':
                    for rr in range(r + 1, num_rows):
                        temp_col = 0
                        for next_cell in table.rows[rr].cells:
                            while (rr, temp_col) in grid:
                                temp_col += 1
                            if temp_col == col:
                                tcPr_next = next_cell._tc.get_or_add_tcPr()
                                vMerge_next = tcPr_next.find(qn('w:vMerge'))
                                if vMerge_next is not None:
                                    vMerge_next_val = vMerge_next.get(qn('w:val'))
                                    if vMerge_next_val is None or vMerge_next_val != 'restart':
                                        rowspan += 1
                                break
                            temp_col += 1

            for i in range(r, r + rowspan):
                for j in range(col, col + colspan):
                    if i == r and j == col:
                        grid[(i, j)] = (cell, rowspan, colspan)
                    else:
                        grid[(i, j)] = None
            col += colspan
            max_cols = max(max_cols, col)

    html = "<table border='1'>"
    for r in range(num_rows):
        html += "<tr>"
        c = 0
        while c < max_cols:
            if (r, c) not in grid:
                c += 1
                continue
            cell_info = grid[(r, c)]
            if cell_info is None:
                c += 1
                continue
            cell, rowspan, colspan = cell_info
            attrs = ""
            if rowspan > 1:
                attrs += f' rowspan="{rowspan}"'
            if colspan > 1:
                attrs += f' colspan="{colspan}"'
            cell_text = cell.text.strip()
            html += f"<td{attrs}>{cell_text}</td>"
            c += colspan
        html += "</tr>"
    html += "</table>"
    return html

def parse_heading_level(text: str, style_name: str) -> Optional[tuple]:
    """
    Парсит текст заголовка и стиль, чтобы определить его уровень и номер.
    Возвращает (уровень, номер_раздела, текст_без_номера).
    Если в тексте номера нет – номер_раздела возвращается как пустая строка.
    """
    text = text.strip()
    section_pattern = r'^(\d+(?:\.\d+)*)\s*\.?(.*)$'
    match = re.match(section_pattern, text)
    if match:
        section_num = match.group(1)
        title = match.group(2).strip()
        if title == "":
            title = text
            section_num = ""
        else:
            level = len(section_num.split('.'))
            return (level, section_num, title)
    style_pattern = r'\d+'
    style_match = re.search(style_pattern, style_name)
    if style_match:
        level = int(style_match.group())
        return (level, "", text)
    return None

def get_parent_section(section_map: Dict[str, str], current_section: str, level: int) -> str:
    """
    Определяет родительский раздел по текущему номеру раздела.
    Если уровень равен 1, это корневой раздел, иначе ищем родителя на предыдущем уровне.
    """
    if level == 1:
        return "(корневой раздел)"
    parts = current_section.split('.')
    if len(parts) > 1:
        parent_num = '.'.join(parts[:-1])
        return section_map.get(parent_num, "(корневой раздел)")
    return "(корневой раздел)"

def extract_sections(doc_path: str) -> List[Dict]:
    """
    Извлекает разделы документа Word с сохранением иерархии, используя стили заголовков.
    Помимо текста, в контент включаются таблицы. Картинки, найденные в абзацах, добавляются
    как отдельные дочерние разделы с указанием родительского раздела, в котором они были обнаружены.
    """
    doc = Document(doc_path)
    sections = []
    section_map = {}
    current_section = None
    current_content = []
    previous_level = 0
    heading_counters = []

    for block in iter_block_items(doc):
        if isinstance(block, Paragraph):
            style_name = block.style.name if block.style and block.style.name else ""
            text, images = get_paragraph_text_and_images(block)
            
            # Если это заголовок
            if "Заголовок" in style_name or "Heading" in style_name or "Title" in style_name:
                # Удаляем теги <img>, если вдруг они остались в тексте
                text = re.sub(r'<img[^>]*>', '', text)
                print(text)
                if current_section is not None:
                    sections.append({
                        "parent": get_parent_section(section_map, current_section, previous_level),
                        "title": section_map[current_section],
                        "content": "\n".join(current_content)
                    })
                current_content = []
                parsed = parse_heading_level(text, style_name)
                if parsed:
                    level, explicit_num, title = parsed
                else:
                    level = 1
                    explicit_num = ""
                    title = text
                if not explicit_num:
                    while len(heading_counters) < level:
                        heading_counters.append(0)
                    heading_counters[level - 1] += 1
                    for i in range(level, len(heading_counters)):
                        heading_counters[i] = 0
                    section_num = ".".join(str(heading_counters[i]) for i in range(level))
                else:
                    section_num = explicit_num
                current_section = section_num
                section_map[current_section] = title
                previous_level = level
                # Если в заголовке обнаружены картинки – добавляем их как отдельные дочерние разделы,
                # указывая название раздела в качестве родителя.
                for idx, img_html in enumerate(images, start=1):
                    sections.append({
                        "parent": section_map[current_section],
                        "title": f"Картинка",
                        "content": img_html
                    })
            else:
                # Обычный абзац – добавляем текст в контент раздела
                if text:
                    current_content.append(text)
                # Если в абзаце есть картинки – каждая добавляется как отдельный дочерний раздел
                for idx, img_html in enumerate(images, start=1):
                    sections.append({
                        "parent": section_map[current_section] if current_section in section_map else "(корневой раздел)",
                        "title": f"Картинка",
                        "content": img_html
                    })
        elif isinstance(block, Table):
            table_html = get_table_html(block)
            if table_html:
                current_content.append("\n" + table_html)
    
    if current_section is not None:
        sections.append({
            "parent": get_parent_section(section_map, current_section, previous_level),
            "title": section_map[current_section],
            "content": "\n".join(current_content)
        })
    
    return sections


def get_excel_images(ws) -> Dict[tuple, List[str]]:
    """
    Извлекает изображения с листа Excel и сопоставляет их с ячейками, в которых они находятся.
    Возвращает словарь, где ключ – (номер_строки, номер_столбца), а значение – список HTML-тегов <img>.
    """
    images = {}
    if hasattr(ws, '_images'):
        for img in ws._images:
            if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                row = img.anchor._from.row + 1
                col = img.anchor._from.col + 1
                try:
                    image_data = img._data()
                except Exception:
                    if hasattr(img, 'path') and os.path.exists(img.path):
                        with open(img.path, 'rb') as f:
                            image_data = f.read()
                    else:
                        continue
                image_base64 = base64.b64encode(image_data).decode('utf-8')
                img_html = f'<img src="data:image/png;base64,{image_base64}" alt="Картинка">'
                if (row, col) in images:
                    images[(row, col)].append(img_html)
                else:
                    images[(row, col)] = [img_html]
    return images

def get_excel_sheet_html(ws, include_images: bool = True) -> str:
    """
    Преобразует лист Excel (worksheet) в HTML‑таблицу с учетом объединённых ячеек.
    Если include_images=True, то изображения НЕ включаются в ячейки (они будут обработаны отдельно).
    """
    images = {}
    if include_images:
        images = get_excel_images(ws)
    max_row = ws.max_row
    max_col = ws.max_column

    merged_info = {}
    merged_cells_set = set()
    for merged_range in ws.merged_cells.ranges:
        min_row, min_col, max_row_range, max_col_range = (merged_range.min_row, merged_range.min_col,
                                                          merged_range.max_row, merged_range.max_col)
        rowspan = max_row_range - min_row + 1
        colspan = max_col_range - min_col + 1
        merged_info[(min_row, min_col)] = (rowspan, colspan)
        for r in range(min_row, max_row_range + 1):
            for c in range(min_col, max_col_range + 1):
                if (r, c) != (min_row, min_col):
                    merged_cells_set.add((r, c))
    
    html = "<table border='1'>"
    for r in range(1, max_row + 1):
        html += "<tr>"
        c = 1
        while c <= max_col:
            if (r, c) in merged_cells_set:
                c += 1
                continue
            cell = ws.cell(row=r, column=c)
            rowspan, colspan = 1, 1
            if (r, c) in merged_info:
                rowspan, colspan = merged_info[(r, c)]
            cell_value = "" if cell.value is None else str(cell.value)
            # Если включено добавление изображений, то изображения не вставляются в ячейку,
            # а будут добавлены как отдельные секции.
            if include_images and (r, c) in images:
                pass
            attrs = ""
            if rowspan > 1:
                attrs += f' rowspan="{rowspan}"'
            if colspan > 1:
                attrs += f' colspan="{colspan}"'
            html += f"<td{attrs}>{cell_value}</td>"
            c += colspan
        html += "</tr>"
    html += "</table>"
    return html

def extract_excel_sections(excel_path: str) -> List[Dict]:
    """
    Обрабатывает Excel‑файл. Для каждого листа создаётся основной раздел,
    где имя листа используется как заголовок, а содержимое листа (HTML‑таблица без встроенных изображений)
    – как контент. Все найденные изображения и встроенные документы добавляются как отдельные дочерние разделы
    с родителем, равным имени листа.
    """
    wb = load_workbook(excel_path, data_only=True)
    sections = []
    for sheet_name in wb.sheetnames:
        print(sheet_name)
        ws = wb[sheet_name]
        images_dict = get_excel_images(ws)
        html_table = get_excel_sheet_html(ws, include_images=False)
        # Основной раздел для листа
        sections.append({
            "parent": "(корневой раздел)",
            "title": sheet_name,
            "content": html_table
        })
        # Разделы для изображений
        for (row, col), img_list in images_dict.items():
            img_title = f"Картинка (ячейка {row}, {col})"
            img_content = "".join(img_list)
            sections.append({
                "parent": sheet_name,
                "title": img_title,
                "content": img_content
            })

    return sections

def save_sections_to_file(sections: List[Dict], output_file: str) -> None:
    """
    Сохраняет извлечённые разделы в текстовый файл.
    """
    with open(output_file, "w", encoding="utf-8") as f:
        for section in sections:
            f.write(f"Заголовок: {section['title']}\n")
            f.write(f"Родитель: {section['parent'] or '(корневой раздел)'}\n")
            f.write(f"Контент: {section['content'][:1000]}...\n")
            f.write("-" * 40 + "\n")

def process_documents(folder_path: str, output_file: str) -> None:
    """
    Обрабатывает все документы Word и Excel в указанной папке и сохраняет результаты в файл.
    """
    all_sections = []
    if not os.path.exists(folder_path):
        raise FileNotFoundError(f"Папка не найдена: {folder_path}")
    
    for filename in os.listdir(folder_path):
        if filename.endswith(".docx"):
            file_path = os.path.join(folder_path, filename)
            try:
                sections = extract_sections(file_path)
                all_sections.extend(sections)
                print(f"Успешно обработан файл: {filename}")
            except Exception as e:
                print(f"Ошибка при обработке файла {filename}: {str(e)}")
        if filename.endswith(".xlsx"):
            file_path = os.path.join(folder_path, filename)
            try:
                excel_sections = extract_excel_sections(file_path)
                all_sections.extend(excel_sections)
                print(f"Успешно обработан Excel‑файл: {filename}")
            except Exception as e:
                print(f"Ошибка при обработке Excel‑файла {filename}: {str(e)}")
    
    save_sections_to_file(all_sections, output_file)
    print(f"Результаты сохранены в {output_file}")

if __name__ == "__main__":
    # Пример обработки Word‑документов и Excel‑файлов
    docs_folder = "docs"
    output_file = "sections_output.txt"
    process_documents(docs_folder, output_file)
